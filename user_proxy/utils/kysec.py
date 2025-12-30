import logging

from openpyxl import load_workbook
from sqlalchemy.orm.attributes import flag_modified

from user_proxy.db import db_session
from user_proxy.models.user import User
from user_proxy.utils.ldap import login_precheck


def parse_user_list_from_excel(excel_file):
    parsed_data = parse_excel(excel_file)
    for user in parsed_data:
        exist_user = login_precheck(user['uid'])
        if exist_user:
            logging.info('user: %s already exist', user['uid'])
            if exist_user.user_data.get('department') == user['department'] and exist_user.user_data.get('username') == user['username']:
                continue
            logging.info('update department from %s to %s, username from %s to %s', exist_user.user_data.get('department'), user['department'], exist_user.user_data['username'], user['username'])
            exist_user.user_data['department'] = user['department']
            exist_user.user_data['username'] = user['username']
            flag_modified(exist_user, 'user_data')
        else:
            logging.info('user: %s, department: %s created!', user['uid'], user['department'])
            user = User.make_user(
                uid=user['uid'], ext_uname=user['uid'],
                department=user['department'], department_id=None, username=user['username']
            )
            db_session.add(user)
    db_session.commit()


def parse_excel(excel_file):
    work_book = load_workbook(excel_file, data_only=True)
    if not work_book.worksheets:
        return []
    sheet = work_book.worksheets[0]
    logging.info(
        "origin table info: title-%s, row_start-%s, row_end-%s, col_start-%s, col_end-%s",
        sheet.title, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column
    )
    parsed_data = []
    for row in range(1, sheet.max_row):
        username = sheet._get_cell(row, 1).value
        uid = sheet._get_cell(row, 2).value
        department = sheet._get_cell(row, 3).value
        parsed_data.append({
            'uid': str(uid),
            'department': department,
            'username': username,
        })
    return parsed_data


if __name__ == '__main__':
    parse_excel('/home/skyrover/下载/人员_1__1__1_.xlsx')
