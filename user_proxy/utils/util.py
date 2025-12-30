# -*-coding:utf-8-*-
from io import BytesIO

import openpyxl

from user_proxy.web_services import UserWebService


def process_users(results):
    fix_results = [['ID', '用户名', '部门', '姓名', '用户角色', '是否来自OA', '用户状态']]
    for item in results:
        role_name = '-'.join([role_info['name'] for role_info in item['roles']])
        oa_user = '是' if item['oa_user'] else '否'
        allow_login = '开启' if item['allow_login'] else '关闭'
        fix_results.append([item['id'], item['ext_uname'], item['user_data'].get('department') or '', item['user_data']['username'] or '', role_name, oa_user, allow_login])

    return fix_results


def save_user_excel(results, path, ret_raw=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.get_sheet_by_name('Sheet'))
    ws = wb.create_sheet('用户列表')
    process_results = process_users(results)
    for row_idx, row_data in enumerate(process_results, 1):
        for col_idx, col_name in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, col_name)
    if ret_raw:
        ret_io = BytesIO()
        wb.save(ret_io)
        ret_io.seek(0)
        return ret_io
    if path:
        wb.save(path)


if __name__ == '__main__':
    users = UserWebService.get_all_users(current_user_id=10)
    save_user_excel(users, '/opt/PycharmProjects/trident/users.xlsx')
