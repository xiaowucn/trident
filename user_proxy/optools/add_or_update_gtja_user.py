# -*-coding:utf-8-*-
import logging
import sys

from openpyxl import load_workbook

from user_proxy.db import db_session
from user_proxy.models.user import Role, User


def bath_update_user_roles(file_path, role_ids=''):
    if not role_ids:
        logging.info("empty role_ids")
        return
    role_ids = role_ids.split(",")
    roles = db_session.query(Role).filter(Role.id.in_(role_ids)).all()
    if not roles:
        logging.info("role: %s not found".format(role_ids))
        return
    user_info = parse_user_excel(file_path)
    logging.info('parse excel res: len: %s, user_info: %s', len(user_info), user_info)
    exist_user_map = dict(db_session.query(User.ext_uname, User).filter(User.ext_uname.in_(user_info)))
    for ext_uname, user_name in user_info.items():
        if ext_uname in exist_user_map:
            user = exist_user_map[ext_uname]
            logging.info('modify user: %s role: %s to role: %s', user.ext_uname, [role.name for role in user.roles], [role.name for role in roles])
        else:
            logging.info('create user, ext_uname: %s, user_name: %s', ext_uname, user_name)
            user = User.make_user(uid=ext_uname, ext_uname=ext_uname, username=user_name, _from='gtja')
        user.roles = roles
    db_session.commit()


def parse_user_excel(file_path):
    work_book = load_workbook(file_path, data_only=True)
    if not work_book.worksheets:
        return []
    sheet = work_book.worksheets[0]
    parsed_data = {str(sheet._get_cell(row, 1).value): str(sheet._get_cell(row, 3).value) for row in range(2, sheet.max_row + 1)}
    return parsed_data


if __name__ == '__main__':
    file_path = sys.argv[1]
    role_ids = sys.argv[2]
    bath_update_user_roles(file_path, role_ids)
