# -*-coding:utf-8-*-
import datetime
import logging
import math
import os
import time
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.sql.functions import count

from user_proxy import config
from user_proxy.db import db_session
from user_proxy.models.user import VisitRecord, User


def utc_to_date_string(utc, fmt='%Y/%m/%d'):
    return time.strftime(fmt, time.localtime(utc))


def utc_from_date_string(date, fmt='%Y/%m/%d'):
    return int(time.mktime(time.strptime(date, fmt)))


def get_sub_system_map():
    sub_sys_map = {'trident': '月均登录人次'}
    sub_sys_info = config.get_config('front_config.sub_sys_config') or {}
    for sub_sys, sys_data in sub_sys_info.items():
        if sys_data['open']:
            sub_sys_map[sub_sys] = f'{sys_data["title"]}模块月均访问人次'
    return sub_sys_map


def init_sub_system_data():
    sub_system_map = get_sub_system_map()
    return {system: 0 for system in sub_system_map}


def get_visit_stat():
    user_count = db_session.query(count(User.id)).filter(User.deleted == 0).scalar()

    current_date = datetime.datetime.today()
    start_utc = utc_from_date_string(str(current_date.year), fmt='%Y')
    trident_visit_records = db_session.query(VisitRecord).filter(VisitRecord.created_utc >= start_utc, VisitRecord.visit_sys == 'trident').all()
    sub_sys_records = dict(
        db_session.query(VisitRecord.visit_sys, count(VisitRecord.id)).filter(VisitRecord.created_utc >= start_utc).group_by(VisitRecord.visit_sys))

    # user_login group by month date
    stat_data = defaultdict(dict)
    for visit_record in trident_visit_records:
        stat_key = utc_to_date_string(visit_record.created_utc, fmt='%Y/%m')
        if stat_key not in stat_data:
            stat_data[stat_key] = {'user_ids': set()}
        stat_data[stat_key]['user_ids'].add(visit_record.user_id)

    # calculate user_login total count
    user_login_count = 0
    for stat_info in stat_data.values():
        user_login_count += len(stat_info['user_ids'])

    # trident and sub_systems total count
    sub_system_data = init_sub_system_data()
    sub_system_data.update(sub_sys_records)

    # fix cn_name and calculate average
    total_month = current_date.month
    sub_system_map = get_sub_system_map()
    cn_sys_data = {}
    for sub_sys, sys_count in sub_system_data.items():
        cn_sys_name = sub_system_map[sub_sys]
        cn_sys_data[cn_sys_name] = math.ceil(sys_count / total_month)

    stat_res = {
        '系统正常状态用户数': user_count,
        '月均登录用户数': math.ceil(user_login_count / total_month),
        **cn_sys_data
    }
    logging.info('访问数据: %s', stat_res)
    file_path = os.path.join(config.project_root, 'icbccs_user_dashboard.xlsx')
    workbook = Workbook()
    sheet = workbook.create_sheet(f'工银瑞信数据统计{datetime.datetime.strftime(current_date, "%Y-%m-%d")}', 0)
    export_data = [list(stat_res.keys()), list(stat_res.values())]
    for row, row_data in enumerate(export_data):
        for col, value in enumerate(row_data):
            sheet.cell(row + 1, col + 1, value)
            sheet.column_dimensions[get_column_letter(col + 1)].width = 30
    workbook.save(file_path)

    return stat_res


if __name__ == '__main__':
    get_visit_stat()
